"""Rellena OTC 62176: código Herinox -> ContPAQi código + especificación."""
from __future__ import annotations

from pathlib import Path

import openpyxl

# Herinox code -> ContPAQi code + specification (from ContPAQi catalog API)
MAPPING: dict[str, dict[str, str | None]] = {
    "CAN015": {
        # Present in ContPAQi; absent from Herinox largos export
        "contpaq": "CAN015",
        "spec": "CANAL A36 6 X 13.0 LB/FT A 20 FT (PESO 260 LB)",
        "note": "mismo codigo ContPAQi; NO existe en Herinox largos",
    },
    "RED027": {
        "contpaq": "RED027",
        "spec": "VARILLA LISA DE ACERO A36 0.75 IN A 20 FT",
        "note": None,
    },
    "ANG022": {
        "contpaq": "ANG022",
        "spec": "ANGULO A36 2 X 2 X 0.25 IN A 20 FT (PESO 61.5 LB)",
        "note": None,
    },
    "PTR016": {
        "contpaq": "PTR016",
        "spec": "PTR A36 THK 0.25 X 4.0 X 4.0 IN X 20 FT (PESO 244 LB)",
        "note": None,
    },
    "HR166": {
        # Herinox: TUBO A36 CED 40 2 IN — no exact ContPAQi SKU for 2.0 IN CED 40
        "contpaq": None,
        "spec": None,
        "note": "sin equivalencia ContPAQi exacta (CED 40 2.0 IN)",
    },
    "HR164": {
        # Herinox: TUBO A36 CED 40 5 IN -> ContPAQi TUB010
        "contpaq": "TUB010",
        "spec": "TUBO A36 CED 40 5.0 IN A 20 FT",
        "note": None,
    },
    "TUB007": {
        "contpaq": "TUB007",
        "spec": "TUBO A36 CED 40 3.0 IN A 20 FT",
        "note": None,
    },
    "SLC051": {
        "contpaq": "SLC051",
        "spec": "SOLERA A36 4.0 X .50 IN A 20 FT (PESO 137 LB)",
        "note": None,
    },
}


def fill_desc(meta: dict[str, str | None]) -> str:
    contpaq = meta.get("contpaq")
    spec = meta.get("spec")
    note = meta.get("note")
    if contpaq and spec:
        return f"{contpaq} | {spec}"
    if note:
        return f"(sin equivalencia) | {note}"
    return "(sin equivalencia) | N/D"


def main() -> None:
    path_in = Path(r"C:\Users\jose_rosales\Downloads\LISTA DE LARGOS OTC 62176.xlsx")
    path_out = Path(r"C:\Users\jose_rosales\Downloads\LISTA DE LARGOS OTC 62176_RESPUESTA.xlsx")
    wb = openpyxl.load_workbook(path_in)
    ws = wb.active
    ws["C2"] = "CODIGO CONTPAQi + ESPECIFICACION:"

    for row in range(3, 13):
        code = str(ws[f"B{row}"].value or "").strip().upper()
        meta = MAPPING.get(code)
        if not meta:
            ws[f"C{row}"] = "(revisar) | sin datos"
            print(code, "MISSING MAP")
            continue
        text = fill_desc(meta)
        ws[f"C{row}"] = text
        print(f"{code} -> {text}")

    wb.save(path_out)
    print("SAVED", path_out)


if __name__ == "__main__":
    main()
