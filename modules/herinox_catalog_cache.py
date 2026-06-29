"""
Respaldo local del catálogo Herinox (precios y dimensiones).

Tras cada consulta exitosa a la BD/API se guarda:
  - cache/herinox_catalog_largos.json   (perfiles LARGO)
  - cache/herinox_plates_snapshot.json  (placas nesting)
  - Plates.xlsx                         (espejo legible / respaldo manual)

Si Herinox no responde, se cargan esos archivos en orden: JSON → XLSX.
"""
from __future__ import annotations

import json
import os
import shutil
from datetime import datetime, timezone
from typing import Any

import config

_CACHE_VERSION = 1
_LARGOS_HEADERS = [
    "Thickness",
    "Material",
    "Arga Code",
    "Length",
    "Width",
    "LB",
    "MXN",
    "$$/LB",
    "Stock",
]
_SHEET_EMPRESA = "STOCK GRUPO ARGA"
_SHEET_PROVEEDOR = "STOCK PROVEEDOR"


def _cache_dir() -> str:
    path = config.ruta_persistente(os.path.join("cache", ""))
    os.makedirs(path, exist_ok=True)
    return path


def _largos_json_path() -> str:
    return os.path.join(_cache_dir(), "herinox_catalog_largos.json")


def _plates_json_path() -> str:
    return os.path.join(_cache_dir(), "herinox_plates_snapshot.json")


def plates_xlsx_path() -> str:
    return config.ruta_persistente("Plates.xlsx")


def _plates_xlsx_seed() -> str:
    return config.ruta_recurso(os.path.join("modules", "Plates.xlsx"))


def _now_iso() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def _atomic_write_json(path: str, payload: dict) -> None:
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    tmp = f"{path}.tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
    os.replace(tmp, path)


def _read_json(path: str) -> dict | None:
    if not os.path.isfile(path):
        return None
    try:
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
        return data if isinstance(data, dict) else None
    except Exception:
        return None


# ---------------------------------------------------------------------------
# LARGOS
# ---------------------------------------------------------------------------


def guardar_snapshot_largos(
    catalogo: list[dict[str, Any]],
    *,
    source: str = "db",
    usd_mxn: float | None = None,
) -> str:
    """Persiste el último catálogo LARGO visto desde Herinox."""
    path = _largos_json_path()
    payload = {
        "version": _CACHE_VERSION,
        "kind": "largos",
        "updated_at": _now_iso(),
        "source": str(source or "db"),
        "usd_mxn": float(usd_mxn) if usd_mxn else None,
        "count": len(catalogo or []),
        "items": list(catalogo or []),
    }
    _atomic_write_json(path, payload)
    return path


def cargar_snapshot_largos() -> tuple[list[dict[str, Any]], dict[str, Any]]:
    data = _read_json(_largos_json_path())
    if not data:
        return [], {}
    items = data.get("items") or []
    if not isinstance(items, list):
        return [], {}
    meta = {
        "updated_at": data.get("updated_at"),
        "source": data.get("source") or "cache_json",
        "usd_mxn": data.get("usd_mxn"),
        "count": len(items),
        "path": _largos_json_path(),
    }
    return [dict(x) for x in items if isinstance(x, dict)], meta


# ---------------------------------------------------------------------------
# PLACAS (nesting)
# ---------------------------------------------------------------------------


def _normalizar_filas_placa(filas: list | None) -> list[list]:
    out: list[list] = []
    for row in filas or []:
        if not isinstance(row, (list, tuple)):
            continue
        out.append([row[i] if i < len(row) else "" for i in range(9)])
    return out


def guardar_snapshot_placas(
    empresa_rows: list,
    proveedor_rows: list,
    *,
    source: str = "db",
    dof_rate: float | None = None,
    dof_source: str | None = None,
) -> dict[str, str]:
    """Persiste inventario de placas (JSON + Plates.xlsx)."""
    emp = _normalizar_filas_placa(empresa_rows)
    prov = _normalizar_filas_placa(proveedor_rows)
    json_path = _plates_json_path()
    payload = {
        "version": _CACHE_VERSION,
        "kind": "plates",
        "updated_at": _now_iso(),
        "source": str(source or "db"),
        "dof_rate": float(dof_rate) if dof_rate else None,
        "dof_source": dof_source,
        "empresa_count": len(emp),
        "proveedor_count": len(prov),
        "empresa": emp,
        "proveedor": prov,
    }
    _atomic_write_json(json_path, payload)
    xlsx_path = exportar_plates_xlsx(emp, prov)
    return {"json": json_path, "xlsx": xlsx_path}


def cargar_snapshot_placas() -> tuple[list[list], list[list], dict[str, Any]]:
    data = _read_json(_plates_json_path())
    if not data:
        return [], [], {}
    emp = _normalizar_filas_placa(data.get("empresa"))
    prov = _normalizar_filas_placa(data.get("proveedor"))
    if not emp and not prov:
        return [], [], {}
    meta = {
        "updated_at": data.get("updated_at"),
        "source": f"cache_json:{data.get('source', 'unknown')}",
        "dof_rate": data.get("dof_rate"),
        "dof_source": data.get("dof_source"),
        "path": _plates_json_path(),
    }
    return emp, prov, meta


def exportar_plates_xlsx(empresa_rows: list, proveedor_rows: list) -> str:
    """Escribe Plates.xlsx junto al ejecutable (formato SHEETS legacy)."""
    try:
        from openpyxl import Workbook
    except ImportError:
        return ""

    dest = plates_xlsx_path()
    os.makedirs(os.path.dirname(dest) or ".", exist_ok=True)
    tmp = f"{dest}.tmp.xlsx"

    wb = Workbook()
    ws_emp = wb.active
    ws_emp.title = _SHEET_EMPRESA
    ws_emp.append(list(_LARGOS_HEADERS))
    for row in _normalizar_filas_placa(empresa_rows):
        ws_emp.append(row)

    ws_prov = wb.create_sheet(_SHEET_PROVEEDOR)
    ws_prov.append(list(_LARGOS_HEADERS))
    for row in _normalizar_filas_placa(proveedor_rows):
        ws_prov.append(row)

    wb.save(tmp)
    os.replace(tmp, dest)
    return dest


def cargar_plates_xlsx(path: str | None = None) -> tuple[list[list], list[list], dict[str, Any]]:
    """Lee Plates.xlsx (persistente o plantilla empaquetada)."""
    candidatos = []
    if path:
        candidatos.append(path)
    candidatos.append(plates_xlsx_path())
    seed = _plates_xlsx_seed()
    if seed and os.path.isfile(seed):
        candidatos.append(seed)

    try:
        from openpyxl import load_workbook
    except ImportError:
        return [], [], {}

    for p in candidatos:
        if not p or not os.path.isfile(p):
            continue
        try:
            wb = load_workbook(p, read_only=True, data_only=True)
            sheets = wb.worksheets
            emp_ws = sheets[0] if sheets else None
            prov_ws = sheets[1] if len(sheets) > 1 else None

            def _rows(ws) -> list[list]:
                if ws is None:
                    return []
                out: list[list] = []
                header_seen = False
                for row in ws.iter_rows(values_only=True):
                    if not row or all(v is None or str(v).strip() == "" for v in row):
                        continue
                    if not header_seen:
                        header_seen = True
                        first = str(row[0] or "").strip().lower()
                        if first in {"thickness", "espesor"}:
                            continue
                    out.append([row[i] if i < len(row) else "" for i in range(9)])
                return out

            emp = _rows(emp_ws)
            prov = _rows(prov_ws)
            wb.close()
            if emp or prov:
                meta = {
                    "updated_at": datetime.fromtimestamp(os.path.getmtime(p), tz=timezone.utc)
                    .replace(microsecond=0)
                    .isoformat(),
                    "source": f"xlsx:{p}",
                    "path": p,
                }
                return emp, prov, meta
        except Exception:
            continue
    return [], [], {}


def cargar_respaldo_placas() -> tuple[list[list], list[list], dict[str, Any]]:
    """Fallback: JSON reciente → Plates.xlsx."""
    emp, prov, meta = cargar_snapshot_placas()
    if emp or prov:
        return emp, prov, meta
    return cargar_plates_xlsx()


def copiar_plates_xlsx_seed_si_falta() -> str | None:
    """Primera ejecución: copia plantilla empaquetada si no hay Plates.xlsx local."""
    dest = plates_xlsx_path()
    if os.path.isfile(dest):
        return dest
    seed = _plates_xlsx_seed()
    if not seed or not os.path.isfile(seed):
        return None
    try:
        os.makedirs(os.path.dirname(dest) or ".", exist_ok=True)
        shutil.copy2(seed, dest)
        return dest
    except Exception:
        return None
